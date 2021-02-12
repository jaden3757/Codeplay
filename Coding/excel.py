# -*- coding: utf-8 -*-

import openpyxl # pip install openpyxl 필수
wb = openpyxl.load_workbook('itemo.xlsx')
wb.save('item.xlsx')
wb = openpyxl.load_workbook('item.xlsx')

# 읽기부분
def getxl(sn, x, y): # sn이라는 이름을 가진 시트의 가로x번째 세로y번째에 관한 값 얻기
    sheet = wb[sn]
    return sheet.cell(column=x, row=y).value

def getxllist(sn, y): # sn이라는 이름을 가진 시트에 y번째 가로줄을 전부 얻기(리스트형태로)
    sheet = wb[sn]
    ls = []
    i = 1
    while i < 20:
        if sheet.cell(column=i, row=y).value != None:
            ls.append(sheet.cell(column=i, row=y).value)
        i += 1
    return ls

# 쓰기부분
def writexl(sn, x, y, txt): # sn이라는 이름을 가진 시트의 xy좌표에 txt값을 지정
    sheet = wb[sn]
    sheet.cell(column=x, row=y).value = txt
    wb.save('item.xlsx')

def writexllist(sn, y, txt):
    sheet = wb[sn]
    i = 1
    while i < 20: # y번째 가로줄 초기화
        sheet.cell(column=i, row=y).value = None
        i += 1
    i = 1

    for text in txt: # y번째 가로줄 바꾸기
        sheet.cell(column=i, row=y).value = text
        i += 1
    wb.save('item.xlsx')

def delxllist(sn, y, val): # sn이름의 시트에서 y번째 가로줄에서 val에 해당하는 값 제거(val는 str과 int모두 가능하지만 int일때는 리스트가 기준이라는 것을 알아둬야함 : 엑셀은 1부터 시작, 파이썬 리스트는 0부터 시작)
    sheet = wb[sn]
    ls = getxllist(sn, y)
    if type(val) == str:
        i = 0
        for text in ls:
            if text == val:
                del ls[i]
            i += 1
    else:
        del ls[val]
    writexllist(sn, y, ls)

def addxllist(sn, y, val): #sn시트의 y번째 가로줄에 val값을 추가
    sheet = wb[sn]
    ls = getxllist(sn, y)
    ls.append(val)
    writexllist(sn, y, ls)

# 전체적인 엑셀 작성법 설명
# -시트
#   시트는 방 하나당 하나씩 지정한다(방 이름이 sample2라면 시트 이름에 sample2라고 그대로 적는 것을 권장)
#   (sp2에 보면 sheetname이라 적힌 곳에 지정된 시트의 이름을 적어줘야함 ex) 위의 'sample2')
#   예외로 메인과 아이템 설명 관련 시트는 따로 있음
#   메인의 1번째 가로줄에는 소지중인 아이템 목록을 적음
# -가로줄
#   각 방마다 1개의 시트를 지정해 줬다면
#   한 가로줄 = 하나의 템 묶음
#   활용 예시 ) box = itemobject('box.png', '박스', width, height, x, y)
#               box.item = [sheetname, 1] # sheetname은 미리 지정해야함 / 1은 1번째 가로줄을 의미
# -아이템의 고유 값과 설명
#   엑셀 파일 어디에서든지 아이템을 추가했을 때 items 시트에 해당 아이템이 없다면 반드시 추가해줘야함
#   items 시트에 보면 한 가로줄에 3개의 칸이 있다
#   1번째 칸 : 아이템의 이름
#   2번째 칸 : 아이템의 고유 값(무게 / 인벤토리에서 차지하는 값)
#   3번째 칸 : 아이템의 설명(반드시 적을 필요는 없다)
#   ::예시:: 
#   |   똥   | 10 | 푸짐한 똥이다 |
#   | 오징어 | 20 | 빨판이 크다   |


#테스트하는 부분
# writexllist('sp2', 1, ['마음']) #1번째 가로줄에 지정
# writexllist('sp2', 2, ['밥', '갑오징어'])
# writexllist('main', 1, ['지도', '손전등', '창고열쇠', '보안키', '연구일지', '오징어 맛살', '똥', '500원'])